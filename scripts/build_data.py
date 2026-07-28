from __future__ import annotations

import json
import math
import re
import urllib.request
import zipfile
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any
import xml.etree.ElementTree as ET

import geopandas as gpd
import pandas as pd
from openpyxl import load_workbook
from pyproj import Transformer
from shapely.geometry import Point
import unicodedata


ROOT = Path(__file__).resolve().parents[1]
PUBLIC_DATA = ROOT / "public" / "data"
PUBLIC_DATA.mkdir(parents=True, exist_ok=True)
AGGREGATES_URL = "https://ftp.ibge.gov.br/Censos/Censo_Demografico_2022/Agregados_por_Setores_Censitarios/Agregados_por_Setor_csv/Agregados_por_setores_basico_BR_20260520.zip"

PE_BOUNDS = {
    "min_lat": -9.7,
    "max_lat": -7.1,
    "min_lng": -42.0,
    "max_lng": -34.7,
}

COMPESA_WORKS_FILE = "Plano de Investimentos - 26jul26 SAESPRI.xlsx"
SDA_WORKS_FILE_PATTERN = "*SDA*.xlsx"
IPA_POCOS_FILE_PATTERN = "Po*.xlsx"
IPA_BARREIROS_FILE = "Barreiros ok.xlsx"
IPA_KML_PATTERN = "*BARREIROS*.kml"
UTM_24S_TO_WGS84 = Transformer.from_crs("EPSG:31984", "EPSG:4326", always_xy=True)
UTM_25S_TO_WGS84 = Transformer.from_crs("EPSG:31985", "EPSG:4326", always_xy=True)


def parse_coord(value: Any, axis: str | None = None) -> float | None:
    if value is None:
        return None
    if isinstance(value, float) and math.isnan(value):
        return None
    text = str(value).strip()
    if not text:
        return None
    normalized = text.replace("#", "").replace("−", "-").replace(",", ".")
    numbers = [float(item) for item in re.findall(r"-?\d+(?:\.\d+)?", normalized)]
    if not numbers:
        return None
    sign = -1 if re.search(r"(^|\s)-|[WSO]|\bOESTE\b|\bSUL\b", normalized.upper()) else 1
    if len(numbers) >= 3:
        value = abs(numbers[0]) + abs(numbers[1]) / 60 + abs(numbers[2]) / 3600
        result = sign * value
    else:
        result = numbers[0]
    if axis == "lat" and result > 0 and 7 <= result <= 10:
        result *= -1
    if axis == "lng" and result > 0 and 34 <= result <= 42:
        result *= -1
    return result


def in_pernambuco(lat: float | None, lng: float | None) -> bool:
    if lat is None or lng is None:
        return False
    return (
        PE_BOUNDS["min_lat"] <= lat <= PE_BOUNDS["max_lat"]
        and PE_BOUNDS["min_lng"] <= lng <= PE_BOUNDS["max_lng"]
    )


def clean_text(value: Any) -> str:
    if value is None or (isinstance(value, float) and math.isnan(value)):
        return ""
    return str(value).strip()


def clean_number(value: Any) -> float:
    if value is None or (isinstance(value, float) and math.isnan(value)):
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = clean_text(value)
    if not text:
        return 0.0
    text = re.sub(r"[^\d,.-]", "", text)
    if "," in text and "." in text:
        text = text.replace(".", "").replace(",", ".")
    elif "," in text:
        text = text.replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return 0.0


def clean_date(value: Any) -> str:
    if value is None or (isinstance(value, float) and math.isnan(value)):
        return ""
    timestamp = pd.to_datetime(value, errors="coerce")
    if pd.isna(timestamp):
        return clean_text(value)
    return timestamp.date().isoformat()


def first_existing_file(candidates: list[Path]) -> Path | None:
    return next((path for path in candidates if path.exists()), None)


def first_matching_file(patterns: list[str], folders: list[Path]) -> Path | None:
    for folder in folders:
        for pattern in patterns:
            matches = sorted(folder.glob(pattern))
            if matches:
                return matches[0]
    return None


def normalize_key(value: Any) -> str:
    text = clean_text(value)
    text = text.replace("Ăş", "ú").replace("Ă­", "í").replace("ĂŁ", "ã").replace("Ă©", "é")
    text = text.replace("Săo", "São").replace("SAO", "SÃO").replace("săo", "são")
    text = unicodedata.normalize("NFD", text)
    text = "".join(ch for ch in text if unicodedata.category(ch) != "Mn")
    text = re.sub(r"[^A-Za-z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip().upper()


def column_by_key(df: pd.DataFrame, *terms: str) -> str | None:
    wanted = [normalize_key(term) for term in terms]
    for column in df.columns:
        key = normalize_key(column)
        if all(term in key for term in wanted):
            return str(column)
    return None


def read_offset_sheet(path: Path, sheet_name: str) -> pd.DataFrame:
    df = pd.read_excel(path, sheet_name=sheet_name, header=3).dropna(how="all").copy()
    df.columns = [clean_text(column) for column in df.columns]
    order_col = column_by_key(df, "ORD")
    if order_col:
        df = df[pd.to_numeric(df[order_col], errors="coerce").notna()].copy()
    return df


def transform_utm_24s(lat_utm: Any, lng_utm: Any) -> tuple[float | None, float | None]:
    northing = clean_number(lat_utm)
    easting = clean_number(lng_utm)
    if not northing or not easting:
        return None, None
    lng, lat = UTM_24S_TO_WGS84.transform(easting, northing)
    if not in_pernambuco(lat, lng):
        return None, None
    return lat, lng


def transform_utm_xy(
    easting_value: Any,
    northing_value: Any,
    municipality: str = "",
    municipality_geometries: dict[str, Any] | None = None,
    preferred_zone: Any = "",
) -> tuple[float | None, float | None]:
    easting = clean_number(easting_value)
    northing = clean_number(northing_value)
    if not easting or not northing:
        return None, None
    zone_key = normalize_key(preferred_zone)
    transformers = [("24S", UTM_24S_TO_WGS84), ("25S", UTM_25S_TO_WGS84)]
    if "25" in zone_key:
        transformers = [("25S", UTM_25S_TO_WGS84), ("24S", UTM_24S_TO_WGS84)]
    elif "24" in zone_key:
        transformers = [("24S", UTM_24S_TO_WGS84), ("25S", UTM_25S_TO_WGS84)]
    candidates = []
    for zone, transformer in transformers:
        lng, lat = transformer.transform(easting, northing)
        if in_pernambuco(lat, lng):
            candidates.append((zone, lat, lng))
    municipality_key = normalize_key(municipality)
    geometry = municipality_geometries.get(municipality_key) if municipality_geometries and municipality_key else None
    if geometry:
        for _, lat, lng in candidates:
            point_geom = Point(lng, lat)
            if geometry.contains(point_geom) or geometry.buffer(0.03).contains(point_geom):
                return lat, lng
    if candidates:
        _, lat, lng = candidates[0]
        return lat, lng
    return None, None


def find_header_row(raw: pd.DataFrame, required: list[str]) -> int | None:
    for index in range(min(25, len(raw))):
        keys = [normalize_key(value) for value in raw.iloc[index].tolist()]
        if all(any(term in key for key in keys) for term in required):
            return index
    return None


def read_table_with_header(path: Path, sheet_name: str, required: list[str]) -> pd.DataFrame:
    raw = pd.read_excel(path, sheet_name=sheet_name, header=None)
    header = find_header_row(raw, required)
    if header is None:
        return pd.DataFrame()
    df = pd.read_excel(path, sheet_name=sheet_name, header=header).dropna(how="all").copy()
    df.columns = [clean_text(column) for column in df.columns]
    first = df.columns[0]
    df = df[pd.to_numeric(df[first], errors="coerce").notna()].copy()
    return df


def column_exact_or_contains(df: pd.DataFrame, exact: str, contains: str | None = None) -> str | None:
    wanted = normalize_key(exact)
    for column in df.columns:
        if normalize_key(column) == wanted:
            return str(column)
    if contains:
        return column_by_key(df, contains)
    return None


def point(
    layer: str,
    name: Any,
    lat: Any,
    lng: Any,
    municipality: Any = "",
    status: Any = "",
    extra: dict[str, Any] | None = None,
) -> dict[str, Any] | None:
    parsed_lat = parse_coord(lat, "lat")
    parsed_lng = parse_coord(lng, "lng")
    if not in_pernambuco(parsed_lat, parsed_lng):
        return None
    return {
        "layer": layer,
        "name": clean_text(name) or "Sem nome informado",
        "municipality": clean_text(municipality),
        "lat": parsed_lat,
        "lng": parsed_lng,
        "status": clean_text(status),
        "extra": {k: clean_text(v) for k, v in (extra or {}).items() if clean_text(v)},
    }


def read_dessalinizadores() -> list[dict[str, Any]]:
    preferred = ROOT / "06.SRHS_DESSALINIZADORES_COMUNIDADES_COORDENADAS.xlsx"
    path = preferred if preferred.exists() else ROOT / "DESSALINIZADORES_COMUNIDADES_COORDENADAS.xlsx"
    df = pd.read_excel(path, header=3)
    rows = []
    for _, row in df.dropna(how="all").iterrows():
        item = point(
            "dessalinizadores",
            row.get("LOCALIDADE"),
            row.get("LATITUDE"),
            row.get("LONGITUDE"),
            row.get("MUNICÍPIO"),
            row.get("DESSALINIZADOR"),
        )
        if item:
            rows.append(item)
    return rows


def read_pocos() -> list[dict[str, Any]]:
    path = ROOT / "POCOS_COMUNIDADES_COORDENADAS.xlsx"
    df = pd.read_excel(path, sheet_name="Plan1", header=1)
    rows = []
    for _, row in df.dropna(how="all").iterrows():
        item = point(
            "pocos",
            row.get("Localidade"),
            row.get("Latitude"),
            row.get("Longitude"),
            row.get("Municipio "),
            row.get("STATUS"),
            {"empresa": row.get("Empresa")},
        )
        if item:
            rows.append(item)
    return rows


def read_sisar() -> list[dict[str, Any]]:
    path = ROOT / "SAA__SISAR_coordenadas.xlsx"
    rows = []
    for sheet in ["Moxotó", "Alto Pajeú"]:
        df = pd.read_excel(path, sheet_name=sheet, header=1)
        for _, row in df.dropna(how="all").iterrows():
            coord = clean_text(row.get("coordenadas"))
            parts = [part.strip() for part in coord.split(",")]
            if len(parts) < 2:
                continue
            item = point(
                "sisar",
                row.get("SISTEMA DE ABASTECIMENTO DE ÁGUA"),
                parts[0],
                parts[1],
                "",
                sheet,
                {"observação": row.get("obs")},
            )
            if item:
                rows.append(item)
    return rows


def read_barragens() -> list[dict[str, Any]]:
    path = ROOT / "LISTA_BARRAGENS_SRHS_SNISB_2026.04.29.csv"
    df = pd.read_csv(path, sep=";", encoding="latin1", header=2)
    rows = []
    for _, row in df.dropna(how="all").iterrows():
        item = point(
            "barragens",
            row.get("Nome_da_Barragem"),
            row.get("Latitude"),
            row.get("Longitude"),
            row.get("Município "),
            row.get("Nível_de_Perigo_Global "),
            {
                "uso": row.get("Uso_Principal"),
                "risco": row.get("Categoria_de_Risco"),
                "dano potencial": row.get("Dano_Potencial_Associado"),
                "capacidade hm3": row.get("Capacidade_hm³ "),
            },
        )
        if item:
            rows.append(item)
    return rows


def read_outorgas(path: Path, layer: str) -> list[dict[str, Any]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = {name: pos for pos, name in enumerate(headers) if name}
    needed = [
        "INT_NU_LATITUDE",
        "INT_NU_LONGITUDE",
        "ING_SG_UFMUNICIPIO",
        "ING_NM_MUNICIPIO",
        "EMP_NM_EMPREENDIMENTO",
        "EMP_NM_USUARIO",
        "INT_TSU_DS",
        "INT_TCH_DS",
        "FIN_TFN_DS",
        "OUT_TP_SITUACAOOUTORGA",
        "INT_QT_VAZAOMEDIA",
    ]
    positions = {name: idx[name] for name in needed if name in idx}
    rows = []
    for values in ws.iter_rows(min_row=2, values_only=True):
        uf = clean_text(values[positions["ING_SG_UFMUNICIPIO"]]) if "ING_SG_UFMUNICIPIO" in positions else ""
        lat = values[positions["INT_NU_LATITUDE"]]
        lng = values[positions["INT_NU_LONGITUDE"]]
        if uf and uf != "PE":
            continue
        item = point(
            layer,
            values[positions.get("EMP_NM_EMPREENDIMENTO", positions.get("EMP_NM_USUARIO", 0))],
            lat,
            lng,
            values[positions.get("ING_NM_MUNICIPIO", 0)],
            values[positions.get("OUT_TP_SITUACAOOUTORGA", 0)],
            {
                "tipo": values[positions.get("INT_TSU_DS", 0)],
                "condição": values[positions.get("INT_TCH_DS", 0)],
                "finalidade": values[positions.get("FIN_TFN_DS", 0)],
                "vazão média": values[positions.get("INT_QT_VAZAOMEDIA", 0)],
            },
        )
        if item:
            rows.append(item)
    wb.close()
    return rows


def read_drought_municipalities() -> set[str]:
    candidates = list(ROOT.glob("Lista de Munic*.csv")) + list(ROOT.glob("Lista de Municípios*.csv"))
    if not candidates:
        return set()
    df = pd.read_csv(candidates[0], encoding="utf-8-sig")
    municipality_col = next((col for col in df.columns if "MUNIC" in normalize_key(col)), df.columns[-1])
    return {normalize_key(value) for value in df[municipality_col].dropna()}


def ensure_basic_aggregates_zip() -> Path:
    target = ROOT / ".cache" / "ibge_agregados" / "Agregados_por_setores_basico_BR_20260520.zip"
    if not target.exists():
        target.parent.mkdir(parents=True, exist_ok=True)
        urllib.request.urlretrieve(AGGREGATES_URL, target)
    return target


def read_sector_population(cd_setores: set[str]) -> pd.DataFrame:
    zip_path = ensure_basic_aggregates_zip()
    chunks = []
    with zipfile.ZipFile(zip_path) as zf:
        name = zf.namelist()[0]
        with zf.open(name) as f:
            reader = pd.read_csv(
                f,
                sep=";",
                encoding="latin1",
                dtype={"CD_SETOR": str, "CD_UF": str, "v0001": str},
                usecols=["CD_SETOR", "CD_UF", "v0001"],
                chunksize=120_000,
            )
            for chunk in reader:
                pe = chunk[(chunk["CD_UF"] == "26") & (chunk["CD_SETOR"].isin(cd_setores))].copy()
                if not pe.empty:
                    pe["population"] = pd.to_numeric(pe["v0001"].str.replace(",", ".", regex=False), errors="coerce").fillna(0).astype(int)
                    chunks.append(pe[["CD_SETOR", "population"]])
    if not chunks:
        return pd.DataFrame(columns=["CD_SETOR", "population"])
    return pd.concat(chunks, ignore_index=True)


def build_rural_geojson() -> tuple[dict[str, Any], dict[str, Any], dict[str, Any], list[str], gpd.GeoDataFrame]:
    zip_path = ROOT / "PE_setores_CD2022.zip"
    extract_dir = ROOT / ".cache" / "ibge_setores"
    shp_path = extract_dir / "PE_setores_CD2022.shp"
    if not shp_path.exists():
        extract_dir.mkdir(parents=True, exist_ok=True)
        with zipfile.ZipFile(zip_path) as zf:
            zf.extractall(extract_dir)

    gdf = gpd.read_file(shp_path)
    rural = gdf[gdf["SITUACAO"].astype(str).str.lower() == "rural"].copy()
    situacao_code_col = "CD_SITUACAO" if "CD_SITUACAO" in rural.columns else "CD_SIT"
    rural["CD_SITUACAO"] = rural[situacao_code_col].astype(str)
    population = read_sector_population(set(rural["CD_SETOR"].astype(str)))
    rural = rural.merge(population, on="CD_SETOR", how="left")
    rural["population"] = rural["population"].fillna(0).astype(int)
    detail_counts = rural["CD_SITUACAO"].value_counts().to_dict()
    rural = rural[
        [
            "CD_SETOR",
            "SITUACAO",
            "CD_SITUACAO",
            "NM_MUN",
            "AREA_KM2",
            "population",
            "geometry",
        ]
    ]
    rural["geometry"] = rural.geometry.simplify(0.0015, preserve_topology=True)
    geojson = json.loads(rural.to_json())
    drought_names = read_drought_municipalities()
    municipalities = gdf[["CD_MUN", "NM_MUN", "geometry"]].copy()
    municipalities["municipio_key"] = municipalities["NM_MUN"].map(normalize_key)
    municipal_polygons = municipalities.dissolve(by=["CD_MUN", "NM_MUN"], as_index=False)
    drought = municipalities[municipalities["municipio_key"].isin(drought_names)].copy()
    drought = drought.dissolve(by=["CD_MUN", "NM_MUN"], as_index=False)
    drought["geometry"] = drought.geometry.simplify(0.002, preserve_topology=True)
    drought_geojson = json.loads(drought[["CD_MUN", "NM_MUN", "geometry"]].to_json())
    matched = set(drought["municipio_key"].unique())
    unmatched = sorted(name for name in drought_names if name not in matched)
    summary = {
        "total_setores": int(len(gdf)),
        "rural_setores": int(len(rural)),
        "rural_area_km2": round(float(rural["AREA_KM2"].fillna(0).sum()), 2),
        "rural_population": int(rural["population"].sum()),
        "rural_agglomerate_population": int(rural[rural["CD_SITUACAO"].isin(["5", "6", "7"])]["population"].sum()),
        "detail_counts": {str(k): int(v) for k, v in detail_counts.items()},
        "detail_population": {str(k): int(v) for k, v in rural.groupby("CD_SITUACAO")["population"].sum().to_dict().items()},
        "drought_municipalities": int(len(drought)),
    }
    return geojson, summary, drought_geojson, unmatched, municipal_polygons


def enrich_missing_municipalities(layers: dict[str, list[dict[str, Any]]], municipal_polygons: gpd.GeoDataFrame) -> int:
    missing = []
    refs = []
    for layer_name, rows in layers.items():
        for index, item in enumerate(rows):
            if not item.get("municipality"):
                missing.append(item)
                refs.append((layer_name, index))
    if not missing:
        return 0

    points = gpd.GeoDataFrame(
        {"ref": range(len(missing))},
        geometry=gpd.points_from_xy([item["lng"] for item in missing], [item["lat"] for item in missing]),
        crs="EPSG:4674",
    )
    polygons = municipal_polygons[["NM_MUN", "geometry"]].copy()
    if polygons.crs is None:
        polygons = polygons.set_crs("EPSG:4674")
    joined = gpd.sjoin(points, polygons, how="left", predicate="within")

    filled = 0
    for _, row in joined.dropna(subset=["NM_MUN"]).iterrows():
        ref = int(row["ref"])
        layer_name, index = refs[ref]
        layers[layer_name][index]["municipality"] = clean_text(row["NM_MUN"]).upper()
        filled += 1
    return filled


def compesa_path() -> Path | None:
    candidates = [
        ROOT / COMPESA_WORKS_FILE,
        Path.home() / "Downloads" / COMPESA_WORKS_FILE,
    ]
    return next((path for path in candidates if path.exists()), None)


def normalized_status(value: Any) -> str:
    key = normalize_key(value)
    if not key:
        return "Não informado"
    groups = {
        "CONCLUIDO": "Concluído",
        "EM ANDAMENTO": "Em andamento",
        "A LICITAR": "A licitar",
        "EM LICITACAO": "Em licitação",
        "A INICIAR": "A iniciar",
        "PROJETO EM ELABORACAO": "Projeto em elaboração",
        "A ELABORAR PROJETO": "A elaborar projeto",
        "A FAZER": "A fazer",
        "A RETOMAR": "A retomar",
    }
    return groups.get(key, clean_text(value).capitalize())


def compesa_status_phase(status: str) -> str:
    key = normalize_key(status)
    if key == "CONCLUIDO":
        return "Concluídas"
    if key in {"EM ANDAMENTO", "A RETOMAR"}:
        return "Em execução"
    if key in {"A LICITAR", "EM LICITACAO", "A INICIAR", "PROJETO EM ELABORACAO", "A ELABORAR PROJETO", "A FAZER"}:
        return "Planejadas"
    return "Não informado"


def match_municipalities(text: Any, municipality_names: dict[str, str]) -> list[str]:
    normalized_text = f" {normalize_key(text)} "
    if not normalized_text.strip():
        return []
    matches = []
    for key, name in sorted(municipality_names.items(), key=lambda item: len(item[0]), reverse=True):
        if re.search(rf"(?<![A-Z0-9]){re.escape(key)}(?![A-Z0-9])", normalized_text):
            matches.append(name)
            normalized_text = re.sub(rf"(?<![A-Z0-9]){re.escape(key)}(?![A-Z0-9])", " ", normalized_text)
    return sorted(set(matches))


def read_compesa_works(municipal_polygons: gpd.GeoDataFrame) -> dict[str, Any]:
    path = compesa_path()
    if not path:
        return {
            "works": [],
            "municipalities": [],
            "map": {"type": "FeatureCollection", "features": []},
            "unmatched_municipality_texts": [],
            "totals": {
                "works": 0,
                "municipalities": 0,
                "value": 0,
                "population": 0,
                "status_counts": {},
                "phase_counts": {},
                "eixo_counts": {},
                "subeixo_counts": {},
            },
        }

    df = pd.read_excel(path, sheet_name="Investimentos")
    municipality_names = {
        normalize_key(name): clean_text(name)
        for name in municipal_polygons["NM_MUN"].dropna().unique()
    }
    municipality_names = {key: value for key, value in municipality_names.items() if key}
    if "SAO CAITANO" in municipality_names:
        municipality_names["SAO CAETANO"] = municipality_names["SAO CAITANO"]
    aggregate: dict[str, dict[str, Any]] = {}
    works = []
    unmatched = []

    for index, row in df.dropna(how="all").iterrows():
        municipalities_original = clean_text(row.get("Municípios Beneficiados"))
        municipalities = match_municipalities(municipalities_original, municipality_names)
        if not municipalities:
            unmatched.append(municipalities_original)
        status = normalized_status(row.get("Status"))
        phase = compesa_status_phase(status)
        value = clean_number(row.get("Valor Divulgado R$"))
        population = int(round(clean_number(row.get("População Beneficiada"))))
        execution = clean_number(row.get("% Execução Realizado"))
        if execution > 1:
            execution = execution / 100
        work = {
            "id": int(index) + 1,
            "name": clean_text(row.get("Nome da Ação")) or "Sem nome informado",
            "description": clean_text(row.get("Descrição")),
            "type": clean_text(row.get("Tipo")),
            "municipalities_original": municipalities_original,
            "municipalities": municipalities,
            "status": status,
            "phase": phase,
            "source": clean_text(row.get("Fonte de Recurso")),
            "population": population,
            "value": round(value, 2),
            "execution": round(execution, 4),
            "start_date": clean_date(row.get("Data de Início")),
            "end_date": clean_date(row.get("Término Previsto")),
            "eixo": clean_text(row.get("Eixo")) or "Não informado",
            "subeixo": clean_text(row.get("Subeixo")) or "Não informado",
        }
        works.append(work)

        if municipalities:
            share = 1 / len(municipalities)
            for municipality in municipalities:
                item = aggregate.setdefault(
                    municipality,
                    {
                        "municipality": municipality,
                        "works_count": 0,
                        "allocated_value": 0.0,
                        "allocated_population": 0.0,
                        "execution_sum": 0.0,
                        "status_counts": Counter(),
                        "phase_counts": Counter(),
                        "eixo_counts": Counter(),
                    },
                )
                item["works_count"] += 1
                item["allocated_value"] += value * share
                item["allocated_population"] += population * share
                item["execution_sum"] += execution
                item["status_counts"][status] += 1
                item["phase_counts"][phase] += 1
                item["eixo_counts"][work["eixo"]] += 1

    municipalities = []
    for item in aggregate.values():
        phase_counts = dict(item["phase_counts"])
        status_counts = dict(item["status_counts"])
        dominant_phase = max(phase_counts, key=phase_counts.get) if phase_counts else "Não informado"
        municipalities.append(
            {
                "municipality": item["municipality"],
                "works_count": int(item["works_count"]),
                "allocated_value": round(float(item["allocated_value"]), 2),
                "allocated_population": int(round(float(item["allocated_population"]))),
                "avg_execution": round(float(item["execution_sum"]) / max(1, item["works_count"]), 4),
                "status_counts": status_counts,
                "phase_counts": phase_counts,
                "eixo_counts": dict(item["eixo_counts"]),
                "dominant_phase": dominant_phase,
            }
        )
    municipalities = sorted(municipalities, key=lambda item: item["allocated_value"], reverse=True)

    agg_by_key = {normalize_key(item["municipality"]): item for item in municipalities}
    compesa_polygons = municipal_polygons.copy()
    compesa_polygons["municipio_key"] = compesa_polygons["NM_MUN"].map(normalize_key)
    compesa_polygons = compesa_polygons[compesa_polygons["municipio_key"].isin(agg_by_key)].copy()
    for field in [
        "works_count",
        "allocated_value",
        "allocated_population",
        "avg_execution",
        "dominant_phase",
        "status_counts",
        "phase_counts",
        "eixo_counts",
    ]:
        compesa_polygons[field] = compesa_polygons["municipio_key"].map(lambda key: agg_by_key[key][field])
    compesa_polygons["geometry"] = compesa_polygons.geometry.simplify(0.002, preserve_topology=True)
    compesa_geojson = json.loads(
        compesa_polygons[
            [
                "CD_MUN",
                "NM_MUN",
                "works_count",
                "allocated_value",
                "allocated_population",
                "avg_execution",
                "dominant_phase",
                "status_counts",
                "phase_counts",
                "eixo_counts",
                "geometry",
            ]
        ].to_json()
    )

    status_counts = Counter(work["status"] for work in works)
    phase_counts = Counter(work["phase"] for work in works)
    eixo_counts = Counter(work["eixo"] for work in works)
    subeixo_counts = Counter(work["subeixo"] for work in works)
    return {
        "works": works,
        "municipalities": municipalities,
        "map": compesa_geojson,
        "unmatched_municipality_texts": sorted(set(item for item in unmatched if item)),
        "totals": {
            "works": int(len(works)),
            "municipalities": int(len(municipalities)),
            "value": round(float(sum(work["value"] for work in works)), 2),
            "population": int(sum(work["population"] for work in works)),
            "status_counts": dict(status_counts),
            "phase_counts": dict(phase_counts),
            "eixo_counts": dict(eixo_counts),
            "subeixo_counts": dict(subeixo_counts),
        },
    }


def sda_path() -> Path | None:
    return first_matching_file(
        [SDA_WORKS_FILE_PATTERN, "DADOS*AGUAS*SDA*.xlsx", "DADOS*ÁGUAS*SDA*.xlsx"],
        [ROOT, Path.home() / "Downloads"],
    )


def sda_status(value: Any, default: str = "A iniciar") -> str:
    key = normalize_key(value)
    if not key:
        return default
    if "ANDAMENTO" in key:
        return "Em andamento"
    if "CONCLUIDO" in key or "ENTREG" in key:
        return "Entregue"
    return clean_text(value)


def add_sda_municipality(
    aggregate: dict[str, dict[str, Any]],
    municipality: str,
    field: str,
    amount: int = 1,
    population: int = 0,
    status: str = "",
) -> None:
    key = normalize_key(municipality)
    if not key:
        return
    item = aggregate.setdefault(
        key,
        {
            "municipality": title_name(municipality),
            "pad": 0,
            "pad_entregue": 0,
            "pad_andamento": 0,
            "pisf": 0,
            "pisf_entregue": 0,
            "pisf_andamento": 0,
            "aguadas": 0,
            "cisternas_total": 0,
            "cisternas_1_agua": 0,
            "cisternas_2_agua": 0,
            "population": 0,
        },
    )
    item[field] += amount
    item["population"] += population
    if field == "pad":
        if normalize_key(status) == "ENTREGUE":
            item["pad_entregue"] += amount
        elif normalize_key(status) == "EM ANDAMENTO":
            item["pad_andamento"] += amount
    if field == "pisf":
        if normalize_key(status) == "ENTREGUE":
            item["pisf_entregue"] += amount
        elif normalize_key(status) == "EM ANDAMENTO":
            item["pisf_andamento"] += amount


def title_name(value: Any) -> str:
    text = clean_text(value)
    return " ".join(part.capitalize() for part in text.lower().split())


def municipal_quantity_geojson(
    municipal_polygons: gpd.GeoDataFrame,
    aggregate: dict[str, dict[str, Any]],
    field: str,
    label: str,
) -> dict[str, Any]:
    selected_keys = {key for key, row in aggregate.items() if row.get(field, 0)}
    if not selected_keys:
        return {"type": "FeatureCollection", "features": []}
    polygons = municipal_polygons.copy()
    polygons["municipio_key"] = polygons["NM_MUN"].map(normalize_key)
    polygons = polygons[polygons["municipio_key"].isin(selected_keys)].copy()
    polygons["quantity"] = polygons["municipio_key"].map(lambda key: int(aggregate[key].get(field, 0)))
    polygons["layer_label"] = label
    polygons["geometry"] = polygons.geometry.simplify(0.002, preserve_topology=True)
    return json.loads(polygons[["CD_MUN", "NM_MUN", "quantity", "layer_label", "geometry"]].to_json())


def read_sda_actions(municipal_polygons: gpd.GeoDataFrame) -> dict[str, Any]:
    path = sda_path()
    empty = {
        "points": {"sda_pad": [], "sda_pisf": []},
        "records": [],
        "municipalities": [],
        "aguadas_map": {"type": "FeatureCollection", "features": []},
        "cisternas_map": {"type": "FeatureCollection", "features": []},
        "totals": {},
    }
    if not path:
        return empty

    aggregate: dict[str, dict[str, Any]] = {}
    records: list[dict[str, Any]] = []
    pad_points: list[dict[str, Any]] = []
    pisf_points: list[dict[str, Any]] = []

    pad = read_offset_sheet(path, "PAD")
    for _, row in pad.iterrows():
        municipality = clean_text(row.get(column_by_key(pad, "MUNICIPIO")))
        locality = clean_text(row.get(column_by_key(pad, "LOCALIDADE")))
        status = sda_status(row.get(column_by_key(pad, "STATUS")))
        population = int(round(clean_number(row.get(column_by_key(pad, "HAB")))))
        item = point(
            "sda_pad",
            locality,
            row.get(column_by_key(pad, "LATITUDE")),
            row.get(column_by_key(pad, "LONGITUDE")),
            municipality,
            status,
            {"programa": "PAD", "população": population},
        )
        if item:
            pad_points.append(item)
        records.append(
            {
                "id": f"pad-{len(records) + 1}",
                "program": "PAD",
                "type": "Dessalinizador SDA",
                "municipality": municipality,
                "locality": locality,
                "status": status,
                "quantity": 1,
                "population": population,
                "lat": item["lat"] if item else None,
                "lng": item["lng"] if item else None,
                "detail": "",
            }
        )
        add_sda_municipality(aggregate, municipality, "pad", 1, population, status)

    pisf = read_offset_sheet(path, "PISF")
    for column in ["ORD", "EIXO", "CAPTACAO", "NOME DO SISTEMA", "STATUS"]:
        source = column_by_key(pisf, column)
        if source:
            pisf[source] = pisf[source].ffill()
    pisf_mun_col = column_by_key(pisf, "MUNICIPIO")
    pisf_loc_col = column_by_key(pisf, "LOCALIDADE")
    pisf_lat_col = column_by_key(pisf, "LATITUDE")
    pisf_lng_col = column_by_key(pisf, "LONGITUDE")
    pisf_status_col = column_by_key(pisf, "STATUS")
    pisf_system_col = column_by_key(pisf, "NOME", "SISTEMA")
    pisf_pop_col = column_by_key(pisf, "HAB")
    pisf_eixo_col = column_by_key(pisf, "EIXO")
    pisf_rows = pisf[pisf_mun_col].notna() & pisf[pisf_loc_col].notna()
    for _, row in pisf[pisf_rows].iterrows():
        municipality = clean_text(row.get(pisf_mun_col)).replace("Pamamirim", "Parnamirim")
        matched = match_municipalities(municipality, {normalize_key(name): clean_text(name) for name in municipal_polygons["NM_MUN"].dropna().unique()})
        if matched:
            municipality = matched[0]
        locality = clean_text(row.get(pisf_loc_col))
        status = sda_status(row.get(pisf_status_col))
        lat, lng = transform_utm_24s(row.get(pisf_lat_col), row.get(pisf_lng_col))
        population = int(round(clean_number(row.get(pisf_pop_col))))
        if lat is not None and lng is not None:
            pisf_points.append(
                {
                    "layer": "sda_pisf",
                    "name": locality,
                    "municipality": municipality,
                    "lat": lat,
                    "lng": lng,
                    "status": status,
                    "extra": {
                        "programa": "PISF",
                        "sistema": clean_text(row.get(pisf_system_col)),
                        "eixo": clean_text(row.get(pisf_eixo_col)),
                        "população": clean_text(population),
                    },
                }
            )
        records.append(
            {
                "id": f"pisf-{len(records) + 1}",
                "program": "PISF",
                "type": "Sistema simplificado",
                "municipality": municipality,
                "locality": locality,
                "status": status,
                "quantity": 1,
                "population": population,
                "lat": lat,
                "lng": lng,
                "detail": clean_text(row.get(pisf_system_col)),
            }
        )
        add_sda_municipality(aggregate, municipality, "pisf", 1, population, status)

    aguadas = read_offset_sheet(path, "Aguadas")
    agu_mun_col = column_by_key(aguadas, "MUNICIPIO")
    agu_qtd_col = column_by_key(aguadas, "QTD")
    agu_region_col = column_by_key(aguadas, "REGIAO")
    for _, row in aguadas.iterrows():
        municipality = clean_text(row.get(agu_mun_col))
        quantity = int(round(clean_number(row.get(agu_qtd_col))))
        if not municipality or not quantity:
            continue
        records.append(
            {
                "id": f"aguadas-{len(records) + 1}",
                "program": "Aguadas",
                "type": "Pequenas barragens/açudes",
                "municipality": municipality,
                "locality": "",
                "status": "A iniciar",
                "quantity": quantity,
                "population": 0,
                "lat": None,
                "lng": None,
                "detail": clean_text(row.get(agu_region_col)),
            }
        )
        add_sda_municipality(aggregate, municipality, "aguadas", quantity)

    cisternas = pd.read_excel(path, sheet_name="Cisternas", header=3).dropna(how="all").copy()
    cisternas.columns = ["ORD", "MUNICIPIO", "QTD_TOTAL", "LATITUDE", "LONGITUDE", "STATUS", "LOTE", "QTD_1_AGUA", "QTD_2_AGUA"]
    cisternas = cisternas[pd.to_numeric(cisternas["ORD"], errors="coerce").notna()].copy()
    for _, row in cisternas.iterrows():
        municipality = clean_text(row.get("MUNICIPIO"))
        total = int(round(clean_number(row.get("QTD_TOTAL"))))
        first = int(round(clean_number(row.get("QTD_1_AGUA"))))
        second = int(round(clean_number(row.get("QTD_2_AGUA"))))
        if not municipality or not total:
            continue
        records.append(
            {
                "id": f"cisternas-{len(records) + 1}",
                "program": "Cisternas",
                "type": "Cisternas",
                "municipality": municipality,
                "locality": "",
                "status": "A iniciar",
                "quantity": total,
                "population": 0,
                "lat": None,
                "lng": None,
                "detail": clean_text(row.get("LOTE")),
                "first_water": first,
                "second_water": second,
            }
        )
        add_sda_municipality(aggregate, municipality, "cisternas_total", total)
        add_sda_municipality(aggregate, municipality, "cisternas_1_agua", first)
        add_sda_municipality(aggregate, municipality, "cisternas_2_agua", second)

    municipalities = sorted(
        [
            {
                **item,
                "total_actions": item["pad"] + item["pisf"] + item["aguadas"] + item["cisternas_total"],
            }
            for item in aggregate.values()
        ],
        key=lambda item: item["total_actions"],
        reverse=True,
    )
    totals = {
        "pad": len(pad_points),
        "pad_records": int(sum(item["program"] == "PAD" for item in records)),
        "pisf_points": len(pisf_points),
        "pisf_records": int(sum(item["program"] == "PISF" for item in records)),
        "aguadas": int(sum(item.get("quantity", 0) for item in records if item["program"] == "Aguadas")),
        "cisternas": int(sum(item.get("quantity", 0) for item in records if item["program"] == "Cisternas")),
        "cisternas_1_agua": int(sum(item.get("first_water", 0) for item in records if item["program"] == "Cisternas")),
        "cisternas_2_agua": int(sum(item.get("second_water", 0) for item in records if item["program"] == "Cisternas")),
        "population": int(sum(item.get("population", 0) for item in records)),
        "municipalities": len(municipalities),
        "status_counts": dict(Counter(item["status"] for item in records)),
        "program_counts": dict(Counter(item["program"] for item in records)),
    }
    return {
        "points": {"sda_pad": pad_points, "sda_pisf": pisf_points},
        "records": records,
        "municipalities": municipalities,
        "aguadas_map": municipal_quantity_geojson(municipal_polygons, aggregate, "aguadas", "Aguadas SDA"),
        "cisternas_map": municipal_quantity_geojson(municipal_polygons, aggregate, "cisternas_total", "Cisternas SDA"),
        "totals": totals,
    }


def ipa_paths() -> dict[str, Path | None]:
    folders = [Path.home() / "Downloads", ROOT]
    return {
        "pocos": first_matching_file(["Poços.xlsx", IPA_POCOS_FILE_PATTERN], folders),
        "barreiros": first_existing_file([ROOT / IPA_BARREIROS_FILE, Path.home() / "Downloads" / IPA_BARREIROS_FILE]),
        "kml": first_matching_file([IPA_KML_PATTERN], folders),
    }


def ipa_status(row: pd.Series, loc_col: str | None, perf_col: str | None, inst_col: str | None, obs_col: str | None) -> str:
    obs = normalize_key(row.get(obs_col)) if obs_col else ""
    if "SECO" in obs or "IMPRODUTIVO" in obs:
        return "Seco/improdutivo"
    if inst_col and clean_text(row.get(inst_col)):
        return "Instalado"
    if perf_col and clean_text(row.get(perf_col)):
        return "Perfurado"
    if loc_col and clean_text(row.get(loc_col)):
        return "Locado"
    return "Sem status"


def read_ipa_pocos(path: Path | None, municipal_polygons: gpd.GeoDataFrame) -> tuple[list[dict[str, Any]], list[dict[str, Any]], dict[str, Any]]:
    if not path:
        return [], [], {}
    municipality_geometries = {
        normalize_key(row["NM_MUN"]): row.geometry
        for _, row in municipal_polygons[["NM_MUN", "geometry"]].iterrows()
    }
    xl = pd.ExcelFile(path)
    points = []
    records = []
    status_counts: Counter[str] = Counter()
    sheet_counts: Counter[str] = Counter()
    mapped = 0
    for sheet in xl.sheet_names:
        if normalize_key(sheet) == "PLANILHA2":
            continue
        df = read_table_with_header(path, sheet, ["MUNICIPIO", "LOCALIDADE"])
        if df.empty:
            continue
        mun_col = column_by_key(df, "MUNICIPIO")
        local_col = column_by_key(df, "LOCALIDADE")
        x_col = column_by_key(df, "COORDENADA", "X") or column_exact_or_contains(df, "Long.", "LONG")
        y_col = column_by_key(df, "COORDENADA", "Y") or column_exact_or_contains(df, "Lat.", "LAT")
        loc_col = column_exact_or_contains(df, "Loc.")
        perf_col = column_exact_or_contains(df, "Perf.") or column_exact_or_contains(df, "PERF.")
        inst_col = column_exact_or_contains(df, "Inst.") or column_exact_or_contains(df, "INST.")
        vaz_col = column_by_key(df, "VAZ")
        std_col = column_by_key(df, "STD")
        zone_col = column_by_key(df, "ZONA", "UTM")
        obs_col = column_by_key(df, "OBS")
        prop_col = column_by_key(df, "PROPRIETARIO")
        for _, row in df.iterrows():
            municipality = title_name(row.get(mun_col))
            locality = clean_text(row.get(local_col))
            status = ipa_status(row, loc_col, perf_col, inst_col, obs_col)
            lat = lng = None
            if x_col and y_col:
                x = clean_number(row.get(x_col))
                y = clean_number(row.get(y_col))
                if abs(x) > 1000 or abs(y) > 1000:
                    lat, lng = transform_utm_xy(x, y, municipality, municipality_geometries, row.get(zone_col) if zone_col else "")
                else:
                    parsed_lat = parse_coord(row.get(y_col), "lat")
                    parsed_lng = parse_coord(row.get(x_col), "lng")
                    if in_pernambuco(parsed_lat, parsed_lng):
                        lat, lng = parsed_lat, parsed_lng
            record = {
                "id": f"ipa-pocos-{len(records) + 1}",
                "program": "Poços IPA",
                "sheet": sheet,
                "municipality": municipality,
                "locality": locality,
                "status": status,
                "lat": lat,
                "lng": lng,
                "flow": clean_number(row.get(vaz_col)) if vaz_col else 0,
                "std": clean_number(row.get(std_col)) if std_col else 0,
                "owner": clean_text(row.get(prop_col)) if prop_col else "",
                "observation": clean_text(row.get(obs_col)) if obs_col else "",
            }
            records.append(record)
            status_counts[status] += 1
            sheet_counts[sheet] += 1
            if lat is not None and lng is not None:
                mapped += 1
                points.append(
                    {
                        "layer": "ipa_pocos",
                        "name": locality or "Poço IPA",
                        "municipality": municipality,
                        "lat": lat,
                        "lng": lng,
                        "status": status,
                        "extra": {
                            "fonte": "IPA",
                            "aba": sheet,
                            "vazão l/h": record["flow"],
                            "STD mg/l": record["std"],
                            "proprietário": record["owner"],
                        },
                    }
                )
    totals = {
        "pocos": len(records),
        "pocos_mapped": mapped,
        "pocos_unmapped": len(records) - mapped,
        "status_counts": dict(status_counts),
        "sheet_counts": dict(sheet_counts),
    }
    return points, records, totals


def read_ipa_barreiros(path: Path | None, municipal_polygons: gpd.GeoDataFrame) -> tuple[list[dict[str, Any]], dict[str, Any], dict[str, Any], dict[str, Any]]:
    if not path:
        return [], {"type": "FeatureCollection", "features": []}, {"type": "FeatureCollection", "features": []}, {}
    raw = pd.read_excel(path, sheet_name="Barreiros e Barragens", header=None)
    rows = raw.iloc[4:].copy()
    rows = rows[pd.to_numeric(rows.iloc[:, 0], errors="coerce").notna()].copy()
    rows = rows.iloc[:, :11]
    rows.columns = ["ord", "region", "municipality", "bar_authorized", "bar_located", "bar_executed", "bar_note", "bpp_authorized", "bpp_located", "bpp_executed", "bpp_note"]
    records = []
    aggregate: dict[str, dict[str, Any]] = {}
    for _, row in rows.iterrows():
        municipality = title_name(row.get("municipality"))
        key = normalize_key(municipality)
        record = {
            "municipality": municipality,
            "region": clean_text(row.get("region")),
            "bar_authorized": int(clean_number(row.get("bar_authorized"))),
            "bar_located": int(clean_number(row.get("bar_located"))),
            "bar_executed": int(clean_number(row.get("bar_executed"))),
            "bpp_authorized": int(clean_number(row.get("bpp_authorized"))),
            "bpp_located": int(clean_number(row.get("bpp_located"))),
            "bpp_executed": int(clean_number(row.get("bpp_executed"))),
        }
        records.append(record)
        aggregate[key] = {**record, "barreiros_executed": record["bar_executed"], "bpp_executed_qty": record["bpp_executed"]}

    def map_for(field: str, label: str) -> dict[str, Any]:
        selected = {key for key, value in aggregate.items() if value.get(field, 0)}
        polygons = municipal_polygons.copy()
        polygons["municipio_key"] = polygons["NM_MUN"].map(normalize_key)
        polygons = polygons[polygons["municipio_key"].isin(selected)].copy()
        if polygons.empty:
            return {"type": "FeatureCollection", "features": []}
        polygons["quantity"] = polygons["municipio_key"].map(lambda key: int(aggregate[key].get(field, 0)))
        polygons["layer_label"] = label
        polygons["geometry"] = polygons.geometry.simplify(0.002, preserve_topology=True)
        return json.loads(polygons[["CD_MUN", "NM_MUN", "quantity", "layer_label", "geometry"]].to_json())

    totals = {
        "bar_authorized": int(sum(item["bar_authorized"] for item in records)),
        "bar_located": int(sum(item["bar_located"] for item in records)),
        "bar_executed": int(sum(item["bar_executed"] for item in records)),
        "bpp_authorized": int(sum(item["bpp_authorized"] for item in records)),
        "bpp_located": int(sum(item["bpp_located"] for item in records)),
        "bpp_executed": int(sum(item["bpp_executed"] for item in records)),
        "municipalities": len(records),
    }
    return records, map_for("barreiros_executed", "Barreiros IPA executados"), map_for("bpp_executed_qty", "Barragens PP IPA executadas"), totals


def read_ipa_kml_points(path: Path | None) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    if not path:
        return [], {}
    root = ET.parse(path).getroot()
    ns = {"k": "http://www.opengis.net/kml/2.2"}
    placemarks = root.findall(".//k:Placemark", ns) or root.findall(".//Placemark")
    points = []
    polygon_count = 0
    for placemark in placemarks:
        name = clean_text(placemark.findtext("k:name", default="", namespaces=ns) or placemark.findtext("name", default=""))
        has_polygon = placemark.find(".//k:Polygon", ns) is not None or placemark.find(".//Polygon") is not None
        if has_polygon:
            polygon_count += 1
        coords = placemark.findtext(".//k:coordinates", default="", namespaces=ns) or placemark.findtext(".//coordinates", default="")
        valid = []
        for token in coords.strip().split():
            parts = token.split(",")
            if len(parts) < 2:
                continue
            try:
                lng = float(parts[0])
                lat = float(parts[1])
            except ValueError:
                continue
            if in_pernambuco(lat, lng):
                valid.append((lat, lng))
        if not valid:
            continue
        lat = sum(item[0] for item in valid) / len(valid)
        lng = sum(item[1] for item in valid) / len(valid)
        points.append(
            {
                "layer": "ipa_barreiros",
                "name": name or "Barreiro IPA",
                "municipality": "",
                "lat": lat,
                "lng": lng,
                "status": "Georreferenciado",
                "extra": {"fonte": "IPA KML", "tipo": "Polígono" if has_polygon else "Ponto"},
            }
        )
    totals = {
        "kml_placemarks": len(placemarks),
        "kml_mapped": len(points),
        "kml_polygons": polygon_count,
    }
    return points, totals


def read_ipa_actions(municipal_polygons: gpd.GeoDataFrame) -> dict[str, Any]:
    paths = ipa_paths()
    pocos_points, pocos_records, pocos_totals = read_ipa_pocos(paths["pocos"], municipal_polygons)
    bar_records, bar_map, bpp_map, bar_totals = read_ipa_barreiros(paths["barreiros"], municipal_polygons)
    kml_points, kml_totals = read_ipa_kml_points(paths["kml"])
    municipalities: dict[str, dict[str, Any]] = {}
    for record in pocos_records:
        key = normalize_key(record["municipality"])
        item = municipalities.setdefault(key, {"municipality": record["municipality"], "pocos": 0, "pocos_instalados": 0, "pocos_perfurados": 0, "barreiros_executed": 0, "bpp_executed": 0})
        item["pocos"] += 1
        if record["status"] == "Instalado":
            item["pocos_instalados"] += 1
        if record["status"] in {"Instalado", "Perfurado"}:
            item["pocos_perfurados"] += 1
    for record in bar_records:
        key = normalize_key(record["municipality"])
        item = municipalities.setdefault(key, {"municipality": record["municipality"], "pocos": 0, "pocos_instalados": 0, "pocos_perfurados": 0, "barreiros_executed": 0, "bpp_executed": 0})
        item["barreiros_executed"] = record["bar_executed"]
        item["bpp_executed"] = record["bpp_executed"]
    rows = sorted(({**value, "total_actions": value["pocos"] + value["barreiros_executed"] + value["bpp_executed"]} for value in municipalities.values()), key=lambda item: item["total_actions"], reverse=True)
    totals = {**pocos_totals, **bar_totals, **kml_totals, "municipalities": len(rows)}
    return {
        "points": {"ipa_pocos": pocos_points, "ipa_barreiros": kml_points},
        "pocos": pocos_records,
        "barreiros": bar_records,
        "municipalities": rows,
        "barreiros_map": bar_map,
        "bpp_map": bpp_map,
        "totals": totals,
    }


def aggregate_by_municipality(points: list[dict[str, Any]]) -> list[dict[str, Any]]:
    counter: dict[str, Counter[str]] = defaultdict(Counter)
    for item in points:
        municipality = item.get("municipality") or "Sem município"
        counter[municipality][item["layer"]] += 1
    rows = []
    for municipality, counts in sorted(counter.items()):
        total = sum(counts.values())
        rows.append({"municipality": municipality, "total": total, "counts": dict(counts)})
    return sorted(rows, key=lambda row: row["total"], reverse=True)


def main() -> None:
    layers = {
        "pocos": read_pocos(),
        "dessalinizadores": read_dessalinizadores(),
        "sisar": read_sisar(),
        "barragens": read_barragens(),
        "outorgas_subterraneas": read_outorgas(
            ROOT / "Outorgas_validas_de_abastecimento_publico___Aguas_subterraneas_17_04_2026_CNAR.xlsx",
            "outorgas_subterraneas",
        ),
        "outorgas_superficiais": read_outorgas(
            ROOT / "Outorgas_validas_de_abastecimento_publico___Aguas_superficiais_17_04_2026_CNARH.xlsx",
            "outorgas_superficiais",
        ),
    }
    rural_geojson, rural_summary, drought_geojson, unmatched_drought, municipal_polygons = build_rural_geojson()
    sda_actions = read_sda_actions(municipal_polygons)
    ipa_actions = read_ipa_actions(municipal_polygons)
    layers.update(sda_actions["points"])
    layers.update(ipa_actions["points"])
    enriched_count = enrich_missing_municipalities(layers, municipal_polygons)
    compesa_works = read_compesa_works(municipal_polygons)
    all_points = [item for rows in layers.values() for item in rows]
    data = {
        "generated_at": pd.Timestamp.now().isoformat(),
        "bounds": PE_BOUNDS,
        "layers": layers,
        "rural": rural_geojson,
        "rural_summary": rural_summary,
        "drought_municipalities": drought_geojson,
        "unmatched_drought_municipalities": unmatched_drought,
        "enriched_municipalities": enriched_count,
        "compesa_works": compesa_works,
        "sda_actions": sda_actions,
        "ipa_actions": ipa_actions,
        "municipalities": aggregate_by_municipality(all_points),
        "totals": {name: len(rows) for name, rows in layers.items()},
        "source_files": [
            "POCOS_COMUNIDADES_COORDENADAS.xlsx",
            "DESSALINIZADORES_COMUNIDADES_COORDENADAS.xlsx",
            "SAA__SISAR_coordenadas.xlsx",
            "LISTA_BARRAGENS_SRHS_SNISB_2026.04.29.csv",
            "Outorgas_validas_de_abastecimento_publico___Aguas_subterraneas_17_04_2026_CNAR.xlsx",
            "Outorgas_validas_de_abastecimento_publico___Aguas_superficiais_17_04_2026_CNARH.xlsx",
            "PE_setores_CD2022.zip",
            "Lista de Municípios - Lista de Municípios.csv",
            "Agregados_por_setores_basico_BR_20260520.zip",
            COMPESA_WORKS_FILE,
            "DADOS ÁGUAS SDA.xlsx",
            "Poços.xlsx",
            IPA_BARREIROS_FILE,
            "RESUMO BARREIROS GERAL 2024 E 2025 _11.08.2025 atualizado R01.csv.kml",
        ],
    }
    output = PUBLIC_DATA / "dashboard.json"
    output.write_text(json.dumps(data, ensure_ascii=False, separators=(",", ":")), encoding="utf-8")
    print(f"Wrote {output}")
    print(json.dumps(data["totals"], ensure_ascii=False, indent=2))
    print(json.dumps(rural_summary, ensure_ascii=False, indent=2))
    print(json.dumps(compesa_works["totals"], ensure_ascii=False, indent=2))
    print(json.dumps(sda_actions["totals"], ensure_ascii=False, indent=2))
    print(json.dumps(ipa_actions["totals"], ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
